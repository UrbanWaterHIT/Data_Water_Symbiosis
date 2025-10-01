import os
import re
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from collections import OrderedDict

# ------------------------- Basic settings -------------------------
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']  # choose a font (optional)
plt.rcParams['axes.unicode_minus'] = False             # fix minus sign rendering

# ========================= Config section =========================
# Data file path (edit to your actual path)
data_path = r""

# Output directories (figures/CSV)
out_dir_fig = r""
out_dir_tab = r""
os.makedirs(out_dir_fig, exist_ok=True)
os.makedirs(out_dir_tab, exist_ok=True)

# Output files (figures)
save_path1 = os.path.join(out_dir_fig, "gini_batches_admin1.svg")
save_path2 = os.path.join(out_dir_fig, "gini_groups_admin1.svg")

# ------------------------- Read data -------------------------
df = pd.read_excel(data_path)

# ========= Column names (updated) =========
dc_pre_col  = "Datacenter pre"
dc_post_col = "Datacenter post"

# (Optional backward-compat shim: if old names exist, rename to new ones)
if dc_pre_col not in df.columns and "DCP1" in df.columns:
    df = df.rename(columns={"DCP1": dc_pre_col})
if dc_post_col not in df.columns and "DCP2" in df.columns:
    df = df.rename(columns={"DCP2": dc_post_col})

# ------------------------- Optional: at GID_2 level, impute by country mean -------------------------
# Comment out the lines below if you don't want this rough imputation
if 'ISO3' not in df.columns:
    raise KeyError("Missing required column: ISO3")

if 'ISO3' in df.columns:
    if 'GDP' in df.columns:
        df['GDP'] = df.groupby('ISO3')['GDP'].transform(lambda x: x.fillna(x.mean()))
    if dc_pre_col in df.columns:
        df[dc_pre_col] = df.groupby('ISO3')[dc_pre_col].transform(lambda x: x.fillna(x.mean()))

# If Datacenter post is missing, fall back to 0 * Datacenter pre (modify if this contradicts your assumptions)
if dc_post_col not in df.columns:
    df[dc_post_col] = np.nan
df[dc_post_col] = df.apply(
    lambda row: row[dc_post_col] if pd.notnull(row[dc_post_col]) else (
        0 * row[dc_pre_col] if pd.notnull(row[dc_pre_col]) else np.nan
    ),
    axis=1
)

# ------------------------- Robustly derive GID_1 from GID_2 -------------------------
# Format conventions:
#   GID_1: "ISO.NUM_vers"   -> exactly 1 dot before the underscore, e.g., "FRA.1_1"
#   GID_2: "ISO.A.B_vers"   -> 2+ dots before the underscore, e.g., "ARE.2.10_1"
def robust_gid1_from_gid2(code: str) -> str:
    if not isinstance(code, str) or '_' not in code:
        return np.nan
    prefix, vers = code.split('_', 1)  # prefix: "FRA.1" / "ARE.2.10" ; vers: "1"
    if prefix.count('.') <= 1:
        # Already in GID_1 shape; return as-is
        return f"{prefix}_{vers}"
    # Otherwise drop the last level: "ARE.2.10" -> "ARE.2"
    return f"{prefix.rsplit('.', 1)[0]}_{vers}"

if 'GID_2' not in df.columns:
    raise KeyError("Missing required column: GID_2")

# (Re)generate GID_1 robustly (avoid regex errors leading to "no data")
df['GID_1'] = df['GID_2'].apply(robust_gid1_from_gid2)

# ------------------------- Helper functions -------------------------
def weighted_mean(series, weights):
    """Population-weighted mean of `series` by `weights`, ignoring NaNs; returns NaN if no valid samples."""
    m = series.notna() & weights.notna() & (weights > 0)
    if m.sum() == 0:
        return np.nan
    return np.average(series[m], weights=weights[m])

def weighted_gini(x, w):
    """
    Population-weighted Gini (Lorenz curve via trapezoids).
    x: indicator (e.g., DC_pre_pc)
    w: weights (e.g., POP_1)
    """
    x = np.asarray(x, dtype=float)
    w = np.asarray(w, dtype=float)
    m = np.isfinite(x) & np.isfinite(w) & (w > 0)
    if m.sum() == 0:
        return 0.0
    x = x[m]; w = w[m]
    if np.all(x == 0):
        return 0.0

    order = np.argsort(x)
    x = x[order]; w = w[order]

    cw = np.cumsum(w)
    cx = np.cumsum(x * w)
    sumw = cw[-1]
    sumxw = cx[-1]  # sum(w*x)

    p = cw / sumw      # cumulative weight share
    L = cx / sumxw     # cumulative indicator share

    p0 = np.r_[0.0, p[:-1]]
    L0 = np.r_[0.0, L[:-1]]
    area = np.sum((L + L0) * (p - p0) / 2.0)

    return float(1.0 - 2.0 * area)

# ------------------------- Aggregate to GID_1 (first-level admin) -------------------------
required_cols = {'ISO3', 'GID_1', 'population_sum', dc_pre_col, dc_post_col}
missing_cols = [c for c in required_cols if c not in df.columns]
if missing_cols:
    raise KeyError(f"Missing required columns: {missing_cols}")

group_cols = ['ISO3', 'GID_1']
df_admin1 = (
    df.groupby(group_cols, as_index=False)
      .apply(lambda g: pd.Series({
          'POP_1':        g['population_sum'].sum(skipna=True),            # total population of the first-level admin unit
          'DC_pre_pc':    weighted_mean(g[dc_pre_col], g['population_sum']),   # pop-weighted per-capita (pre)
          'DC_post_pc':   weighted_mean(g[dc_post_col], g['population_sum']),  # pop-weighted per-capita (post)
      }))
      .reset_index(drop=True)
)
# Note: Do NOT fill *_pc with 0 at the GID_1 level (it would inflate inequality). Keeping NaN is more robust.

# ------------------------- Country list (for output order) -------------------------
countries = [
    'ESP'
]
# (Removed: any HKG/TWN removal and CHN insertion logic)

# ------------------------- Compute population-weighted Gini by country (based on GID_1) -------------------------
results = []
for country in countries:
    a1 = df_admin1[(df_admin1['ISO3'] == country)]
    if a1.empty:
        print(f"No data: {country}")
        continue

    gini_pre  = weighted_gini(a1['DC_pre_pc'].values,  a1['POP_1'].values)
    gini_post = weighted_gini(a1['DC_post_pc'].values, a1['POP_1'].values)
    gini_diff = gini_post - gini_pre

    results.append({
        'Country':   country,
        'Gini_Pre':  gini_pre,
        'Gini_Post': gini_post,
        'Gini_Diff': gini_diff
    })

df_results = pd.DataFrame(results)
df_results = df_results.replace([np.inf, -np.inf], np.nan).dropna(subset=['Gini_Pre','Gini_Post'])
print("Per-country (GID_1, population-weighted) Gini coefficients and differences:")
print(df_results)

# ------------------------- Increased/Decreased/Unchanged lists and export -------------------------
df_increase  = df_results[df_results['Gini_Diff'] >  0].sort_values('Gini_Diff', ascending=False)
df_decrease  = df_results[df_results['Gini_Diff'] <  0].sort_values('Gini_Diff', ascending=True)
df_unchanged = df_results[np.isclose(df_results['Gini_Diff'], 0, atol=1e-12)]

print("\n[Countries with increased Gini (descending)]")
print(df_increase[['Country','Gini_Pre','Gini_Post','Gini_Diff']])
print("\n[Countries with decreased Gini (ascending)]")
print(df_decrease[['Country','Gini_Pre','Gini_Post','Gini_Diff']])
print("\n[Countries with ~unchanged Gini (|diff|≈0)]")
print(df_unchanged[['Country','Gini_Pre','Gini_Post','Gini_Diff']])

print("\nStats:")
print(f"Increase: {len(df_increase)}; Decrease: {len(df_decrease)}; Unchanged: {len(df_unchanged)}; Total: {len(df_results)}.")

# Export CSVs
df_increase.to_csv(os.path.join(out_dir_tab, "gini_increase_admin1_popweighted.csv"), index=False, encoding="utf-8-sig")
df_decrease.to_csv(os.path.join(out_dir_tab, "gini_decrease_admin1_popweighted.csv"), index=False, encoding="utf-8-sig")
df_unchanged.to_csv(os.path.join(out_dir_tab, "gini_unchanged_admin1_popweighted.csv"), index=False, encoding="utf-8-sig")
df_results.to_csv(os.path.join(out_dir_tab, "gini_summary_admin1_popweighted.csv"), index=False, encoding="utf-8-sig")

# ------------------------- Diagnose "no data" reasons and export -------------------------
expected = set(countries)
present_raw = set(df['ISO3'].dropna().unique().tolist())
present_admin1 = set(df_admin1['ISO3'].dropna().unique().tolist())

missing_in_raw     = sorted(list(expected - present_raw))                     # not in the raw table
missing_in_admin1  = sorted(list((expected & present_raw) - present_admin1))  # present in raw, absent after aggregation
ok_in_admin1       = sorted(list(expected & present_admin1))

diag_rows = []

# Case A: country not present in the raw table
for iso in missing_in_raw:
    diag_rows.append(OrderedDict(ISO3=iso, reason="No records for this country in the raw data (ISO3 missing/mismatch)",
                                 n_gid2=0, n_gid1=0, n_pop_pos=0,
                                 n_dc_pre_valid=0, n_dc_post_valid=0))

# Case B: present in raw data but missing after aggregation
for iso in missing_in_admin1:
    sub = df[df['ISO3'] == iso]
    n_gid2 = len(sub)
    n_gid1 = sub['GID_1'].notna().sum() if 'GID_1' in sub.columns else 0
    n_pop_pos = (sub['population_sum'].fillna(0) > 0).sum()
    n_dc_pre_valid  = sub[dc_pre_col].notna().sum()  if dc_pre_col  in sub.columns else 0
    n_dc_post_valid = sub[dc_post_col].notna().sum() if dc_post_col in sub.columns else 0

    if n_gid2 == 0:
        reason = "Abnormal raw data (should not happen)"
    elif n_gid1 == 0:
        reason = "GID_2→GID_1 rule mismatch caused aggregation failure (check function and GID_2 format)"
    elif n_pop_pos == 0:
        reason = "Population ≤ 0 or missing, invalid weights"
    elif n_dc_pre_valid == 0 and n_dc_post_valid == 0:
        reason = "Datacenter pre/post entirely missing at GID_2 level (pre-imputation didn't cover)"
    else:
        reason = "Other (spot-check GID_2/version, field formats)"

    diag_rows.append(OrderedDict(ISO3=iso, reason=reason,
                                 n_gid2=n_gid2, n_gid1=n_gid1, n_pop_pos=n_pop_pos,
                                 n_dc_pre_valid=n_dc_pre_valid, n_dc_post_valid=n_dc_post_valid))

# Case C: present after aggregation, with quality notes
for iso in ok_in_admin1:
    a1 = df_admin1[df_admin1['ISO3'] == iso]
    n_rows = len(a1)
    n_pop_pos = (a1['POP_1'].fillna(0) > 0).sum()
    n_dc_pre_valid  = a1['DC_pre_pc'].notna().sum()
    n_dc_post_valid = a1['DC_post_pc'].notna().sum()
    reason = "OK (computable)"
    if n_rows <= 1:
        reason = "Only 1 first-level admin unit: Gini = 0 by construction; inequality not captured"
    elif n_pop_pos == 0:
        reason = "Population weights invalid (POP_1 ≤ 0)"
    elif n_dc_pre_valid == 0 or n_dc_post_valid == 0:
        reason = "Per-capita values for one scenario all missing; results unstable (check pre-imputation)"

    diag_rows.append(OrderedDict(ISO3=iso, reason=reason,
                                 n_gid2=np.nan, n_gid1=n_rows, n_pop_pos=n_pop_pos,
                                 n_dc_pre_valid=n_dc_pre_valid, n_dc_post_valid=n_dc_post_valid))

df_diag = pd.DataFrame(diag_rows).sort_values(['reason','ISO3']).reset_index(drop=True)
print("\n[Diagnostics table for missing data and quality]")
print(df_diag)

# Export diagnostics table
df_diag.to_csv(os.path.join(out_dir_tab, "gini_admin1_popweighted_diagnostics.csv"),
               index=False, encoding="utf-8-sig")

# ------------------------- Plot/export only countries with "changes" (two rows) -------------------------
# Tolerance to determine "changed"
tol = 1e-6

# Keep only changed (exclude |diff| <= tol)
df_increase  = df_results[df_results['Gini_Diff'] >  tol].sort_values('Gini_Diff', ascending=False)
df_decrease  = df_results[df_results['Gini_Diff'] < -tol].sort_values('Gini_Diff', ascending=True)
df_changed   = pd.concat([df_increase, df_decrease], ignore_index=True)

n_inc, n_dec, n_tot = len(df_increase), len(df_decrease), len(df_changed)
print("\nOnly countries with changes:")
print(f"Increase: {n_inc}; Decrease: {n_dec}; Total: {n_tot}.")

if n_tot == 0:
    raise ValueError("No countries meet the change threshold (tol). Lower `tol` or check the data.")

# Export only the "changed" results
df_increase.to_csv(os.path.join(out_dir_tab, f"gini_increase_admin1_popweighted_tol{tol}.csv"),
                   index=False, encoding="utf-8-sig")
df_decrease.to_csv(os.path.join(out_dir_tab, f"gini_decrease_admin1_popweighted_tol{tol}.csv"),
                   index=False, encoding="utf-8-sig")
df_changed.to_csv(os.path.join(out_dir_tab,  f"gini_changed_summary_admin1_popweighted_tol{tol}.csv"),
                  index=False, encoding="utf-8-sig")

# ------------------------- Plot: two rows (top: increase; bottom: decrease) -------------------------
save_path_changed = os.path.join(out_dir_fig, "gini_changed_admin1_two_rows.svg")

colors_pre  = "#A76779"  # Pre
colors_post = "#7090B3"  # Post

# Compute figure size based on the max bars in a single row
max_bars = max(n_inc, n_dec, 1)
fig_w = max(16, min(36, 0.6 * max_bars))  # adaptive width
fig_h = 12
fig, axes = plt.subplots(2, 1, figsize=(fig_w, fig_h), sharex=False)

def plot_row(ax, data, title_en, rotate=45, fontsize=14):
    if data.empty:
        ax.text(0.5, 0.5, "No country meets the condition", ha='center', va='center', fontsize=16)
        ax.axis('off')
        return
    x = np.arange(len(data))
    width = 0.38
    ax.bar(x - width / 2, data['Gini_Pre'],  width, label='Pre-Co (GID_1, pop-weighted)',  color=colors_pre)
    ax.bar(x + width / 2, data['Gini_Post'], width, label='Post-Co (GID_1, pop-weighted)', color=colors_post)
    ax.set_xticks(x)
    ax.set_xticklabels(data['Country'], rotation=rotate, ha='right', fontsize=fontsize)
    ax.set_ylabel("Gini coefficient", fontsize=16)
    ax.set_title(title_en, fontsize=18, pad=10)
    ax.legend(fontsize=12)
    ax.grid(axis='y', linestyle='--', alpha=0.2)

# Top row: Gini increased (high to low)
plot_row(axes[0], df_increase, "Gini increased (high to low)", rotate=45, fontsize=12)
# Bottom row: Gini decreased (low to high; more negative on the left)
plot_row(axes[1], df_decrease, "Gini decreased (low to high; more negative on the left)", rotate=45, fontsize=12)

plt.tight_layout(rect=[0, 0, 1, 0.98])
fig.savefig(save_path_changed, format='svg', transparent=True)
plt.show()

print(f"\nSaved SVG: {save_path_changed}")

# ====== Export all-country Gini to Excel (3 columns only) ======
# Note: the Country column equals ISO3; here we rename it to ISO and order columns as required.
# ====== Export all-country Gini to Excel (ISO, Gini pre, Gini Post) ======
import os

# 1) Select and rename columns
df_xlsx = (
    df_results[['Country', 'Gini_Pre', 'Gini_Post']]
      .rename(columns={'Country': 'ISO', 'Gini_Pre': 'Gini pre', 'Gini_Post': 'Gini Post'})
)

# 2) Ensure output directory exists
os.makedirs(out_dir_tab, exist_ok=True)

excel_path = os.path.join(out_dir_tab, "gini_admin1_popweighted_all.xlsx")

# 3) Prefer openpyxl for xlsx; if not installed, fall back to CSV
try:
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        df_xlsx.to_excel(writer, sheet_name="Gini_All", index=False)

    # Optional: set B and C columns to 3 decimal places (requires openpyxl)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb["Gini_All"]
        for col in ("B", "C"):
            for cell in ws[col][1:]:  # skip header
                cell.number_format = "0.000"
        wb.save(excel_path)
    except Exception:
        pass

    print(f"Exported Excel: {excel_path}")

except ModuleNotFoundError:
    # Neither openpyxl nor xlsxwriter available; export CSV instead
    csv_path = excel_path.replace(".xlsx", ".csv")
    df_xlsx.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"Excel engine not detected (openpyxl/xlsxwriter). Exported CSV instead: {csv_path}")
